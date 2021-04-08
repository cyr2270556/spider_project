#
import openpyxl

workbook1 = openpyxl.load_workbook(r'C:\Users\admin\Desktop\车险资源3.xlsx')
sheet = workbook1['Sheet1']

workbook2 = openpyxl.Workbook()
sheet2 = workbook2.create_sheet('chexian')


for i in range(1,sheet.max_row+1):
    name = sheet.cell(i,6).value
    phone = sheet.cell(i, 8).value
    info = sheet.cell(i, 16).value
    sheet2.cell(i,1,name)
    sheet2.cell(i, 2, phone)
    sheet2.cell(i, 3, info)
workbook2.save(r'C:\Users\admin\Desktop\车险资源(检测).xlsx')