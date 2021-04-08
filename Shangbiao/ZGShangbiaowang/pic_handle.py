# 2021/1/11 弃用




# # 图片识别并写入excel模块
# # 将“下载”文件夹中的图片按照数字顺序获取
# # 将获取图片切割成小块图片进行图文转换
# # 转换出的文字再进行辨识度提高处理
# # 将图片数据转换成文字数据导入excel
#
# from PIL import Image as image_P
# import pytesseract
# import cv2
# import openpyxl
# import os
#
#
# def rec_pic(target_root):
#     """
#
#     :param target_root:目标图像识别文件路径
#     :return: 图像识别内容
#     """
#     text = pytesseract.image_to_string((image_P.open(target_root)), lang='chi_sim')
#
#     return text
#
#
# def cut_pic(root):
#     """
#     ！可能需要微调
#     一个图片上有两个数据，把两个数据切出来，up是前一个数据，down是后一个数据
#     :param target_root: 目标图片路径
#     :param save_root:  存储图片路径
#     :return:
#     """
#     img = cv2.imread(root)
#     cropped_up = img[820:880, 200:800]
#     cropped_down = img[1370:1430, 200:800]
#     cv2.imwrite(r'G:\after\up.jpg', cropped_up)
#     cv2.imwrite(r'G:\after\down.jpg', cropped_down)
#     return print('图片切割成功')
#
#
# ###主代码1 ：将一个文件夹中的所有图片做切割，print测试识别出的内容
# # pic_list = os.listdir(r"C:\Users\admin\Downloads")
#
# wb = openpyxl.load_workbook(r'G:\shangbiao_data-1677.xlsx')
# # ws = wb.active
# ws = wb['shangbiao_data']
#
# ###向excel中写入撤销和答辩种类
#
# # 撤销复审决定书
# rec_list1 = ['发文菩暂菖= 撤锏复市决定书', '发文茎鲤= 撤销复市决定书', '发文莒鲤= 撒淌氯轲决定书',]
#
# # 关于撤销连续三年未使用商标的决定
# rec_list2 = ['发文菩暂菖= 关丁撤销连穸] _′′^', '发文鄄鲤= 关 丁 撤销连续三年未使用商标的决定', '发文茎鲤= 关丁提供注册商标使用证据的通知',
#              '发文类型妻 关 丁 撤销连续三年未使用商标的决定', '发文菩暂菖= 关] 鹫放宁衔j堇仝其一年牙〈狂〔…商本盂「拆J决定']
# # 商标撤销复审答辩通知书
# rec_list3 = ['发文菩暂菖= 商标撤销复市答辩通知书', '发文茎鲤= 闹称兄效百 l答辩通知书', '发文茎鲤= 商标撤销复市答辩通知书',
#              '发文菩鲤= 阉怀撤钠复市答辩通知书', '发文类型蕃 商标撤销复市答辩通知书']
# # 商标无效宣告答辩通知书
# rec_list4 = ['发文菩暂菖= 商标无效山懵答辩通知书', '发文茎鲤= 商标无效山懵答辩通知书',
#              '发文鄄鲤= 商标无效山懵答辩通知书','发文类型= 商称兄效画 】答辩通知书']
#
# # 连续三年不使用撤销申请的结案通知
# rec_list5 = ['发文类型= 违续三年不使用撤销 申请的结案通知']
#
# for i in range(1, 4249):
#     if not os.path.exists(r'C:\Users\admin\Downloads\{}.jpg'.format(i)):
#         continue
#     pic_root = r'C:\Users\admin\Downloads\{}.jpg'.format(i)
#     print(pic_root)
#     cut_pic(pic_root)
#     up = rec_pic(r'G:\after\up.jpg')
#     down = rec_pic(r'G:\after\down.jpg')
#     print(up)
#     print(down)
#
#     up_translate = 'error'
#     if up in rec_list1:
#         up_translate = '撤销复审决定书'
#         ws.cell(i*2, 4, up_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2)
#     elif up in rec_list2:
#         up_translate = '关于撤销连续三年未使用商标的决定'
#         ws.cell(i*2, 4, up_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2)
#     elif up in rec_list3:
#         up_translate = '商标撤销复审答辩通知书'
#         ws.cell(i*2, 4, up_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2)
#     elif up in rec_list4:
#         up_translate = '商标无效宣告答辩通知书'
#         ws.cell(i*2, 4, up_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2)
#     elif up in rec_list5:
#         up_translate = '连续三年不使用撤销申请的结案通知'
#         ws.cell(i*2, 4, up_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2)
#
#     down_translate = 'error'
#     if down in rec_list1:
#         down_translate = '撤销复审决定书'
#         ws.cell(i*2+1, 4, down_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2+1)
#     elif down in rec_list2:
#         down_translate = '关于撤销连续三年未使用商标的决定'
#         ws.cell(i*2+1, 4, down_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2+1)
#     elif down in rec_list3:
#         down_translate = '商标撤销复审答辩通知书'
#         ws.cell(i*2+1, 4, down_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2+1)
#     elif down in rec_list4:
#         down_translate = '商标无效宣告答辩通知书'
#         ws.cell(i*2+1, 4, down_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2+1)
#     elif down in rec_list5:
#         down_translate = '连续三年不使用撤销申请的结案通知'
#         ws.cell(i*2+1, 4, down_translate)
#         wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
#         print(i*2+1)
#     os.remove(r'G:\after\up.jpg')
#     os.remove(r'G:\after\down.jpg')

# 发文菩暂菖= 关丁撤销连穸] _′′^
# 发文鄄鲤= 关 丁 撤销连续三年未使用商标的决定
# 发文茎鲤= 撤销复市决定书
# 发文菩暂菖= 商标撤销复市答辩通知书
# 发文菩暂菖= 商标无效山懵答辩通知书
# 发文菩暂菖= 撤锏复市决定书
# 发文茎鲤= 闹称兄效百 l答辩通知书
# 发文菩暂菖= 商标无效山懵答辩通知书
# 发文茎鲤= 闹称兄效百 l答辩通知书


# count 开始为1，up写在第二行，down写在第三行 if执行完count+2 变成3，up写在第四行，down写在第五行，count+2 变成5，up写6，down写7


# print(house_client.estimated_document_count())  # 数据数量

# 图片识别并写入excel模块
# 将“下载”文件夹中的图片按照数字顺序获取
# 将获取图片切割成小块图片进行图文转换
# 转换出的文字再进行辨识度提高处理
# 将图片数据转换成文字数据导入excel

from PIL import Image as image_P
import pytesseract
import cv2
import openpyxl
import os


def rec_pic(target_root):
    """

    :param target_root:目标图像识别文件路径
    :return: 图像识别内容
    """
    text = pytesseract.image_to_string((image_P.open(target_root)), lang='chi_sim')

    return text


def cut_pic(root):
    """
    ！可能需要微调
    一个图片上有两个数据，把两个数据切出来，up是前一个数据，down是后一个数据
    :param target_root: 目标图片路径
    :param save_root:  存储图片路径
    :return:
    """
    img = cv2.imread(root)
    cropped_up = img[820:880, 200:800]
    cropped_down = img[1370:1430, 200:800]
    cv2.imwrite(r'G:\after\up.jpg', cropped_up)
    cv2.imwrite(r'G:\after\down.jpg', cropped_down)
    return print('图片切割成功')


###主代码1 ：将一个文件夹中的所有图片做切割，print测试识别出的内容
# pic_list = os.listdir(r"C:\Users\admin\Downloads")

wb = openpyxl.load_workbook(r'G:\shangbiao_data-1677.xlsx')
# ws = wb.active
ws = wb['shangbiao_data']

###向excel中写入撤销和答辩种类

# 撤销复审决定书
rec_list1 = ['发文菩暂菖= 撤锏复市决定书', '发文茎鲤= 撤销复市决定书', '发文莒鲤= 撒淌氯轲决定书',]

# 关于撤销连续三年未使用商标的决定
rec_list2 = ['发文菩暂菖= 关丁撤销连穸] _′′^', '发文鄄鲤= 关 丁 撤销连续三年未使用商标的决定', '发文茎鲤= 关丁提供注册商标使用证据的通知',
             '发文类型妻 关 丁 撤销连续三年未使用商标的决定', '发文菩暂菖= 关] 鹫放宁衔j堇仝其一年牙〈狂〔…商本盂「拆J决定']
# 商标撤销复审答辩通知书
rec_list3 = ['发文菩暂菖= 商标撤销复市答辩通知书', '发文茎鲤= 闹称兄效百 l答辩通知书', '发文茎鲤= 商标撤销复市答辩通知书',
             '发文菩鲤= 阉怀撤钠复市答辩通知书', '发文类型蕃 商标撤销复市答辩通知书']
# 商标无效宣告答辩通知书
rec_list4 = ['发文菩暂菖= 商标无效山懵答辩通知书', '发文茎鲤= 商标无效山懵答辩通知书',
             '发文鄄鲤= 商标无效山懵答辩通知书','发文类型= 商称兄效画 】答辩通知书']

# 连续三年不使用撤销申请的结案通知
rec_list5 = ['发文类型= 违续三年不使用撤销 申请的结案通知']

for i in range(1, 4249):
    if not os.path.exists(r'C:\Users\admin\Downloads\{}.jpg'.format(i)):
        continue
    pic_root = r'C:\Users\admin\Downloads\{}.jpg'.format(i)
    print(pic_root)
    cut_pic(pic_root)
    up = rec_pic(r'G:\after\up.jpg')
    down = rec_pic(r'G:\after\down.jpg')
    print(up)
    print(down)

    up_translate = 'error'
    if up in rec_list1:
        up_translate = '撤销复审决定书'
        ws.cell(i*2, 4, up_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2)
    elif up in rec_list2:
        up_translate = '关于撤销连续三年未使用商标的决定'
        ws.cell(i*2, 4, up_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2)
    elif up in rec_list3:
        up_translate = '商标撤销复审答辩通知书'
        ws.cell(i*2, 4, up_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2)
    elif up in rec_list4:
        up_translate = '商标无效宣告答辩通知书'
        ws.cell(i*2, 4, up_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2)
    elif up in rec_list5:
        up_translate = '连续三年不使用撤销申请的结案通知'
        ws.cell(i*2, 4, up_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2)

    down_translate = 'error'
    if down in rec_list1:
        down_translate = '撤销复审决定书'
        ws.cell(i*2+1, 4, down_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2+1)
    elif down in rec_list2:
        down_translate = '关于撤销连续三年未使用商标的决定'
        ws.cell(i*2+1, 4, down_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2+1)
    elif down in rec_list3:
        down_translate = '商标撤销复审答辩通知书'
        ws.cell(i*2+1, 4, down_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2+1)
    elif down in rec_list4:
        down_translate = '商标无效宣告答辩通知书'
        ws.cell(i*2+1, 4, down_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2+1)
    elif down in rec_list5:
        down_translate = '连续三年不使用撤销申请的结案通知'
        ws.cell(i*2+1, 4, down_translate)
        wb.save(r'G:\c_spider_work_project\Shangbiao\shangbiao_data-1677.xlsx')
        print(i*2+1)
    os.remove(r'G:\after\up.jpg')
    os.remove(r'G:\after\down.jpg')