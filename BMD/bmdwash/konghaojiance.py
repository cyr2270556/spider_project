#空号检测文件
#将excel中号码提取出，按行放入txt文件上传刀鱼进行空号检测
#将空号检测完4个txt文件回写excel

import openpyxl

workbook = openpyxl.load_workbook(r'C:\Users\admin\Desktop\车险资源2.xlsx')
sheet = workbook['Sheet1']

#空值移位
# for i in range(1,sheet.max_row+1):
#
#     if sheet.cell(i,8).value == None:
#         phone = sheet.cell(i,7).value
#         sheet.cell(i,8,phone)

# workbook.save(r'C:\Users\admin\Desktop\车险资源2.xlsx')
#max_column 列数

#将excel 导入txt
# with open(r'C:\Users\admin\Desktop\车险资源.txt',"w") as f:
#     for i in range(1,sheet.max_row+1):
#         phone = sheet.cell(i,8).value
#         f.write(str(phone)+'\n')


# 读取检测后端txt 重新放入excel
root1 = r'G:\c_spider_work_project\车险资源\沉默号.txt'
root2 = r'G:\c_spider_work_project\车险资源\风险号.txt'
root3 = r'G:\c_spider_work_project\车险资源\活跃号(实号).txt'
root4 = r'G:\c_spider_work_project\车险资源\空号.txt'



chengmo_list=[]
fengxian_list=[]
huoyue_list =[]
konghao_list =[]
with open(root1,"r",encoding='utf-8') as f:
    for line in f:
        chengmo_list.append(line.strip())
with open(root2,"r",encoding='utf-8') as f:
    for line in f:
        fengxian_list.append(line.strip())
with open(root3,"r",encoding='utf-8') as f:
    for line in f:
        huoyue_list.append(line.strip())
with open(root4,"r",encoding='utf-8') as f:
    for line in f:
        konghao_list.append(line.strip())

for i in range(1,sheet.max_row+1):
    phone = sheet.cell(i,8).value
    if str(phone) in chengmo_list:
        print('cm')
        sheet.cell(i,16,'沉默号')
for i in range(1,sheet.max_row+1):
    phone = sheet.cell(i,8).value
    if str(phone) in fengxian_list:
        print('cm')
        sheet.cell(i,16,'风险号')
for i in range(1,sheet.max_row+1):
    phone = sheet.cell(i,8).value
    if str(phone) in huoyue_list:
        print('cm')
        sheet.cell(i,16,'活跃号(实号)')

for i in range(1,sheet.max_row+1):
    phone = sheet.cell(i,8).value
    if str(phone) in konghao_list:
        print('cm')
        sheet.cell(i,16,'空号')
        

workbook.save(r'C:\Users\admin\Desktop\车险资源3.xlsx')

