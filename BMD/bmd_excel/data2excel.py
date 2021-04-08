#2021/1/11

#文字数据导入excel 模块
#先于图片导入excel执行
#x轴字段，y轴id，id按照爬取顺序


import openpyxl
import pymongo

house_client = pymongo.MongoClient("mongodb://localhost:27017/")["cuiworkdb"]["BMD20201224"]

workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('qiye_data')
basic_dict = house_client.find_one()
col_basic_num = 1
basic_list = []
for k in basic_dict:
    sheet.cell(1, col_basic_num, k)
    basic_list.append(k)
    col_basic_num += 1
print(basic_list)
data_row = 2
for line in house_client.find():
    for col in range(1, len(basic_list)+1):
        try:
            sheet.cell(data_row, col, line[basic_list[col-1]])
            if line[basic_list[col-1]] == "":
                sheet.cell(data_row, col, "None")
        except:
            continue
    data_row += 1

workbook.save(r"G:/BMD20201224.xlsx")