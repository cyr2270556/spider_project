import xlrd
import xlwt
from TOOLS.mongosave import MongoDB
import openpyxl
# 10.2.1.122:17017

import pymongo

house_client = pymongo.MongoClient("localhost:27017")["BMD"]["jqr"]

workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('jqr_data')
# print(house_client.estimated_document_count())  # 数据数量

# for line in house_client.find():
#
#     sheet.write(row_n,col_n,value)

basic_dict = house_client.find_one()
col_basic_num = 1
basic_list = []
for k in basic_dict:
    sheet.cell(1, col_basic_num, k)
    basic_list.append(k)
    col_basic_num += 1
print(basic_list)
data_row = 2
for line in house_client.find():

    for col in range(1, len(basic_list)+1):
        print(line)
        print(line[basic_list[2]])
        sheet.cell(data_row, col, line[basic_list[col-1]])
    data_row += 1
    print(data_row)


# {'_id': '82b83a4be172976f0dbe6f94393b6991', 'room_area': '945', 'room_num': '1102', 'advance_sale': '10792', 'room_unit': '1', 'room_floor': '11', 'room_buyer': '邱永志', 'phone': '18615761631', 'building_num': '1', 'entry_name': '主角商厦'}


workbook.save(r'G:\c_spider_work_project\Lbl_personal_work\jqr_data.xls')
